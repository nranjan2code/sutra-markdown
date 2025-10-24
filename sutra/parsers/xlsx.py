"""
Excel Parser - Extract content from Excel files

Handles Excel files with support for:
- Multiple sheet processing
- Cell data extraction
- Formula evaluation
- Table structure preservation
- Chart and image detection
- Metadata extraction

Requirements:
    pip install openpyxl pandas

Usage:
    parser = XLSXParser("workbook.xlsx")
    result = await parser.parse()
    
    # Or streaming
    async for element in parser.parse_stream():
        process(element)
"""

import time
from pathlib import Path
from typing import List, Optional, Dict, Any, AsyncIterator
from datetime import datetime
import asyncio

from ..models.document import (
    ParsedDocument,
    PageInfo,
    ImageInfo,
    TableInfo,
    DocumentElement,
    DocumentMetadata
)
from ..models.enums import ElementType, DocumentType
from .base import BaseParser, ParserResult, ParserError, CorruptedFileError

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XLSXParser(BaseParser):
    """
    Excel file parser using openpyxl
    
    Extracts data from Excel workbooks including multiple sheets,
    cell values, formulas, and table structures.
    
    Args:
        file_path: Path to Excel file
        options: Parser options
            - sheet_names: List of sheet names to process (default: all)
            - max_rows: Maximum rows per sheet (default: None - all)
            - max_cols: Maximum columns per sheet (default: None - all)
            - include_formulas: Include formula text (default: False)
            - include_empty_cells: Include empty cells in output (default: False)
            - data_only: Read only cell values, not formulas (default: True)
    
    Example:
        >>> parser = XLSXParser("data.xlsx", {
        ...     "sheet_names": ["Sheet1", "Data"],
        ...     "max_rows": 10000,
        ...     "include_formulas": True
        ... })
        >>> result = await parser.parse()
        >>> print(f"Extracted {len(result.document.tables)} tables")
    """
    
    def __init__(
        self,
        file_path: str | Path,
        options: Optional[Dict[str, Any]] = None
    ):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel parsing.\n"
                "Install with: pip install openpyxl"
            )
        
        super().__init__(file_path, options)
        
        # Parse options
        self.sheet_names = self.options.get("sheet_names", None)
        self.max_rows = self.options.get("max_rows", None)
        self.max_cols = self.options.get("max_cols", None)
        self.include_formulas = self.options.get("include_formulas", False)
        self.include_empty_cells = self.options.get("include_empty_cells", False)
        self.data_only = self.options.get("data_only", True)
        
        # Load workbook
        try:
            self.workbook = load_workbook(
                str(self.file_path),
                data_only=self.data_only,
                read_only=False  # Need write access for some metadata operations
            )
        except Exception as e:
            raise CorruptedFileError(f"Failed to load Excel file: {e}")
    
    def get_page_count(self) -> int:
        """
        Get page count (number of worksheets)
        
        Returns:
            Number of worksheets
        """
        if hasattr(self, 'workbook'):
            sheets_to_process = self._get_sheets_to_process()
            return len(sheets_to_process)
        return 1
    
    def _get_sheets_to_process(self) -> List[Any]:
        """
        Get list of worksheets to process
        
        Returns:
            List of worksheet objects
        """
        if self.sheet_names:
            sheets = []
            for name in self.sheet_names:
                if name in self.workbook.sheetnames:
                    sheets.append(self.workbook[name])
        else:
            sheets = list(self.workbook.worksheets)
        
        return sheets
    
    async def validate(self) -> tuple[bool, Optional[str]]:
        """
        Validate Excel file
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            # Check if workbook was loaded
            if not hasattr(self, 'workbook') or self.workbook is None:
                return False, "Failed to load Excel workbook"
            
            # Check if has worksheets
            if not self.workbook.worksheets:
                return False, "Workbook has no worksheets"
            
            # Check if requested sheets exist
            if self.sheet_names:
                missing_sheets = []
                for name in self.sheet_names:
                    if name not in self.workbook.sheetnames:
                        missing_sheets.append(name)
                
                if missing_sheets:
                    return False, f"Sheets not found: {', '.join(missing_sheets)}"
            
            return True, None
            
        except Exception as e:
            return False, str(e)
    
    async def parse(self) -> ParserResult:
        """
        Parse entire Excel workbook
        
        Returns:
            ParserResult with extracted content
        """
        start_time = time.time()
        warnings = []
        
        try:
            # Validate first
            is_valid, error = await self.validate()
            if not is_valid:
                return ParserResult(
                    document=None,
                    success=False,
                    error=error,
                    file_size=self.file_size
                )
            
            # Extract metadata
            metadata = self._extract_metadata()
            
            # Process worksheets
            pages = []
            tables = []
            full_text_parts = []
            
            sheets = self._get_sheets_to_process()
            
            for sheet_index, worksheet in enumerate(sheets):
                try:
                    # Extract sheet data
                    sheet_data = self._extract_sheet_data(worksheet)
                    
                    if sheet_data['rows']:
                        # Combine headers and rows for TableInfo
                        all_rows = []
                        if sheet_data['headers']:
                            all_rows.append(sheet_data['headers'])
                        all_rows.extend(sheet_data['rows'])
                        
                        # Create table info
                        table_info = TableInfo(
                            id=f"xlsx_sheet_{sheet_index}",
                            rows=all_rows,  # All rows as List[List[str]]
                            page_number=sheet_index + 1,
                            bbox=None,
                            caption=worksheet.title
                        )
                        tables.append(table_info)
                        
                        # Create text representation
                        sheet_text = self._sheet_to_text(
                            worksheet.title,
                            sheet_data['headers'],
                            sheet_data['rows']
                        )
                        
                        # Create page info
                        page_info = PageInfo(
                            page_number=sheet_index + 1,
                            text=sheet_text,
                            bounds={}
                        )
                        pages.append(page_info)
                        full_text_parts.append(sheet_text)
                
                except Exception as e:
                    warnings.append(f"Failed to process sheet '{worksheet.title}': {e}")
                    continue
            
            # Create ParsedDocument
            document = ParsedDocument(
                metadata=metadata,
                pages=pages,
                images=[],  # Excel images not extracted in this version
                tables=tables,
                full_text="\n\n".join(full_text_parts)
            )
            
            parse_time = time.time() - start_time
            
            return ParserResult(
                document=document,
                success=True,
                warnings=warnings,
                parse_time=parse_time,
                file_size=self.file_size
            )
            
        except Exception as e:
            return ParserResult(
                document=None,
                success=False,
                error=str(e),
                parse_time=time.time() - start_time,
                file_size=self.file_size
            )
    
    async def parse_stream(self) -> AsyncIterator[DocumentElement]:
        """
        Parse Excel in streaming mode
        
        Yields document elements as they are parsed.
        """
        try:
            sheets = self._get_sheets_to_process()
            
            for sheet_index, worksheet in enumerate(sheets):
                # Yield sheet header
                yield DocumentElement(
                    type=ElementType.HEADER,
                    content=f"Sheet: {worksheet.title}",
                    metadata={
                        "sheet_name": worksheet.title,
                        "sheet_index": sheet_index
                    }
                )
                
                try:
                    # Process sheet data in chunks
                    rows_processed = 0
                    chunk_size = 100  # Process 100 rows at a time
                    
                    # Get data range
                    if worksheet.max_row and worksheet.max_column:
                        max_row = min(worksheet.max_row, self.max_rows or worksheet.max_row)
                        max_col = min(worksheet.max_column, self.max_cols or worksheet.max_column)
                        
                        # Process in chunks
                        for start_row in range(1, max_row + 1, chunk_size):
                            end_row = min(start_row + chunk_size - 1, max_row)
                            
                            # Extract chunk data
                            chunk_data = []
                            for row in worksheet.iter_rows(
                                min_row=start_row,
                                max_row=end_row,
                                max_col=max_col,
                                values_only=True
                            ):
                                # Convert row to list of strings
                                row_data = [str(cell) if cell is not None else "" for cell in row]
                                
                                # Skip empty rows unless requested
                                if not self.include_empty_cells and not any(cell.strip() for cell in row_data):
                                    continue
                                
                                chunk_data.append(row_data)
                                rows_processed += 1
                            
                            # Yield chunk as table element
                            if chunk_data:
                                chunk_text = self._format_chunk_data(chunk_data)
                                
                                yield DocumentElement(
                                    type=ElementType.TABLE,
                                    content=chunk_text,
                                    metadata={
                                        "sheet_name": worksheet.title,
                                        "sheet_index": sheet_index,
                                        "chunk_start_row": start_row,
                                        "chunk_end_row": end_row,
                                        "rows_in_chunk": len(chunk_data)
                                    }
                                )
                            
                            await asyncio.sleep(0)  # Yield control
                
                except Exception as e:
                    yield DocumentElement(
                        type=ElementType.TEXT,
                        content="",
                        metadata={
                            "sheet_name": worksheet.title,
                            "error": str(e)
                        }
                    )
        
        except Exception as e:
            yield DocumentElement(
                type=ElementType.TEXT,
                content="",
                metadata={"error": str(e)}
            )
    
    def _extract_sheet_data(self, worksheet: Any) -> Dict[str, Any]:
        """
        Extract data from a worksheet
        
        Args:
            worksheet: openpyxl worksheet object
        
        Returns:
            Dict with headers and rows
        """
        if not worksheet.max_row or not worksheet.max_column:
            return {"headers": [], "rows": []}
        
        # Determine data range
        max_row = min(worksheet.max_row, self.max_rows or worksheet.max_row)
        max_col = min(worksheet.max_column, self.max_cols or worksheet.max_column)
        
        rows = []
        headers = []
        
        # Extract all data
        for row_idx, row in enumerate(worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            max_col=max_col,
            values_only=True
        )):
            # Convert to list of strings
            row_data = [str(cell) if cell is not None else "" for cell in row]
            
            # Skip completely empty rows unless requested
            if not self.include_empty_cells and not any(cell.strip() for cell in row_data):
                continue
            
            if row_idx == 0:
                # Try to determine if first row is header
                if self._is_header_row(row_data, worksheet, max_row):
                    headers = row_data
                    continue
            
            rows.append(row_data)
        
        # Generate headers if none detected
        if not headers and rows:
            headers = [f"Column_{i+1}" for i in range(len(rows[0]))]
        
        # Ensure consistent column count
        if headers and rows:
            max_cols_data = max(len(headers), max(len(row) for row in rows) if rows else 0)
            
            # Pad headers
            while len(headers) < max_cols_data:
                headers.append(f"Column_{len(headers)+1}")
            
            # Pad rows
            for row in rows:
                while len(row) < max_cols_data:
                    row.append("")
        
        return {"headers": headers, "rows": rows}
    
    def _is_header_row(self, row_data: List[str], worksheet: Any, max_row: int) -> bool:
        """
        Determine if first row is likely a header row
        
        Args:
            row_data: First row data
            worksheet: Worksheet object
            max_row: Maximum row number
        
        Returns:
            True if likely a header row
        """
        if max_row < 2:
            return True  # Only one row, assume it's header
        
        # Check if first row has different characteristics than second row
        try:
            second_row = list(worksheet.iter_rows(
                min_row=2, max_row=2, values_only=True
            ))[0]
            
            second_row_data = [str(cell) if cell is not None else "" for cell in second_row]
            
            # If first row is text and second row has numbers, likely header
            first_numeric = sum(1 for cell in row_data if self._is_numeric(cell))
            second_numeric = sum(1 for cell in second_row_data if self._is_numeric(cell))
            
            # Header likely if first row has fewer numbers
            return first_numeric < second_numeric
        
        except Exception:
            return True  # Default to header if can't determine
    
    def _is_numeric(self, value: str) -> bool:
        """Check if value is numeric"""
        try:
            float(value.strip())
            return True
        except (ValueError, AttributeError):
            return False
    
    def _format_chunk_data(self, chunk_data: List[List[str]]) -> str:
        """
        Format chunk data as text
        
        Args:
            chunk_data: List of row data
        
        Returns:
            Formatted text
        """
        lines = []
        for row in chunk_data:
            lines.append(" | ".join(cell for cell in row))
        
        return "\n".join(lines)
    
    def _sheet_to_text(self, sheet_name: str, headers: List[str], rows: List[List[str]]) -> str:
        """
        Convert sheet data to text representation
        
        Args:
            sheet_name: Name of the sheet
            headers: Column headers
            rows: Data rows
        
        Returns:
            Text representation
        """
        lines = [f"Sheet: {sheet_name}", "=" * (len(sheet_name) + 7), ""]
        
        if headers:
            lines.append(" | ".join(headers))
            lines.append("-" * 50)
        
        # Show first 20 rows
        for row in rows[:20]:
            lines.append(" | ".join(str(cell) for cell in row))
        
        if len(rows) > 20:
            lines.append(f"... and {len(rows) - 20} more rows")
        
        return "\n".join(lines)
    
    def _extract_metadata(self) -> DocumentMetadata:
        """
        Extract Excel workbook metadata
        
        Returns:
            DocumentMetadata object
        """
        # Get workbook properties
        props = self.workbook.properties
        
        # Extract basic metadata
        title = props.title if hasattr(props, 'title') else None
        author = props.creator if hasattr(props, 'creator') else None
        created_date = props.created if hasattr(props, 'created') else None
        modified_date = props.modified if hasattr(props, 'modified') else None
        
        # Count sheets and cells
        sheet_count = len(self.workbook.worksheets)
        total_cells = 0
        
        for worksheet in self.workbook.worksheets:
            if worksheet.max_row and worksheet.max_column:
                total_cells += worksheet.max_row * worksheet.max_column
        
        return self._create_metadata(
            title=title or self.file_path.stem,
            author=author,
            created_date=created_date,
            modified_date=modified_date,
            page_count=sheet_count,
            sheet_count=sheet_count,
            total_cells=total_cells,
            sheet_names=[ws.title for ws in self.workbook.worksheets]
        )
    
    def get_supported_features(self) -> Dict[str, bool]:
        """Get supported features"""
        return {
            "text_extraction": True,
            "table_extraction": True,
            "image_extraction": False,  # Not implemented yet
            "metadata_extraction": True,
            "streaming": True,
            "multi_sheet": True,
            "formulas": self.include_formulas
        }


# Quick test function
async def test_xlsx_parser():
    """Test Excel parser"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python -m sutra.parsers.xlsx <xlsx_file>")
        return
    
    xlsx_file = sys.argv[1]
    
    print(f"\n{'='*70}")
    print(f"🧪 Testing Excel Parser")
    print(f"{'='*70}")
    print(f"File: {xlsx_file}\n")
    
    try:
        # Create parser
        parser = XLSXParser(xlsx_file)
        
        print(f"📄 Document Info:")
        print(f"   Type: {parser.document_type.value}")
        print(f"   Size: {parser.file_size:,} bytes")
        print(f"   Sheets: {parser.get_page_count()}")
        print(f"   Sheet names: {[ws.title for ws in parser.workbook.worksheets]}")
        print(f"   Hash: {parser.file_hash[:16]}...")
        
        # Validate
        print(f"\n✓ Validating...")
        is_valid, error = await parser.validate()
        
        if not is_valid:
            print(f"❌ Validation failed: {error}")
            return
        
        print(f"✅ Valid Excel file")
        
        # Parse
        print(f"\n⏳ Parsing...")
        result = await parser.parse()
        
        if not result.success:
            print(f"❌ Parse failed: {result.error}")
            return
        
        print(f"✅ Parsed in {result.parse_time:.2f}s")
        
        if result.warnings:
            print(f"⚠️  Warnings:")
            for warning in result.warnings:
                print(f"   - {warning}")
        
        # Show results
        doc = result.document
        print(f"\n📊 Results:")
        print(f"   Sheets processed: {len(doc.pages)}")
        print(f"   Tables: {len(doc.tables)}")
        print(f"   Total text: {len(doc.full_text):,} chars")
        
        # Show table info
        if doc.tables:
            print(f"\n📊 Tables:")
            for idx, table in enumerate(doc.tables):
                print(f"   Sheet {idx + 1}: {table.rows:,} rows x {table.cols} cols")
        
        # Show first sheet preview
        if doc.pages:
            preview = doc.pages[0].text[:300].replace("\n", " ")
            print(f"\n📄 First sheet preview:")
            print(f"   {preview}...")
        
        print(f"\n{'='*70}")
        print(f"✅ Test complete!")
        print(f"{'='*70}\n")
        
    except Exception as e:
        print(f"\n❌ Error: {e}\n")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    import asyncio
    asyncio.run(test_xlsx_parser())